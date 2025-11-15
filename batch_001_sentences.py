#!/usr/bin/env python3
"""
LLM-Generated Sentences for Batch 1 (Rows 2-31)
High-quality, contextually appropriate sentences for English language learners
"""

# Format: row_number: [sentence1, sentence2, sentence3]
sentences = {
    2: [  # the
        "The teacher explained the lesson clearly to all students.",
        "I saw the most beautiful sunset from the top of the mountain.",
        "Please pass me the book that's sitting on the table."
    ],
    3: [  # of
        "The signing of the treaty brought peace to both nations.",
        "She drank a glass of cold water after her long run.",
        "The museum displays paintings of famous artists from the Renaissance."
    ],
    4: [  # and
        "Sarah bought apples and oranges at the farmer's market.",
        "The movie was long and entertaining, keeping everyone engaged.",
        "He studied hard and passed his final exams with excellent grades."
    ],
    5: [  # to
        "I need to finish my homework before dinner tonight.",
        "She walked to the library to return her borrowed books.",
        "The key to success is dedication and consistent effort."
    ],
    6: [  # in
        "The children are playing in the backyard right now.",
        "She was born in 1995 in a small coastal town.",
        "There's a lot of truth in what you just said."
    ],
    7: [  # that
        "I believe that honesty is always the best policy.",
        "The house that we visited yesterday has been sold.",
        "She was so tired that she fell asleep immediately."
    ],
    8: [  # was
        "The concert was absolutely amazing last night.",
        "He was reading a book when the phone rang.",
        "The weather was perfect for our outdoor wedding."
    ],
    9: [  # his
        "Michael forgot his wallet at home this morning.",
        "The artist is known for his incredible attention to detail.",
        "He shared his sandwich with his younger brother."
    ],
    10: [  # he
        "He arrived early to prepare for the important meeting.",
        "When he heard the news, he couldn't believe it.",
        "Everyone knows that he is the most reliable person on the team."
    ],
    11: [  # it
        "The package arrived yesterday, but it was damaged during shipping.",
        "It takes courage to admit when you're wrong.",
        "I bought a new phone because it has a better camera."
    ],
    12: [  # with
        "She decorated the cake with fresh strawberries and cream.",
        "I completely agree with your assessment of the situation.",
        "He walked home with his best friend after school."
    ],
    13: [  # for
        "Thank you for helping me with this difficult project.",
        "This gift is for you to celebrate your graduation.",
        "She has worked for the same company for fifteen years."
    ],
    14: [  # as
        "As the sun set, the sky turned brilliant shades of orange.",
        "She works as a nurse at the children's hospital.",
        "The situation is not as bad as it first appeared."
    ],
    15: [  # be
        "You should be proud of everything you've accomplished.",
        "The meeting will be held in the conference room tomorrow.",
        "I want to be a doctor when I finish my education."
    ],
    16: [  # had
        "She had already eaten dinner before her friends arrived.",
        "They had never seen snow until they moved to Canada.",
        "If I had known about the party, I would have come."
    ],
    17: [  # you
        "You look wonderful in that new dress.",
        "Can you help me carry these heavy boxes upstairs?",
        "I hope you have a fantastic time on your vacation."
    ],
    18: [  # not
        "I'm not sure if I can attend the meeting tomorrow.",
        "The store is not open on Sundays or public holidays.",
        "She does not like spicy food at all."
    ],
    19: [  # her
        "Jessica brought her laptop to the coffee shop to work.",
        "The news made her incredibly happy and excited.",
        "I gave her my phone number so we could stay in touch."
    ],
    20: [  # on
        "The picture is hanging on the wall above the fireplace.",
        "We're planning a big celebration on her birthday next month.",
        "Please turn on the lights because the room is too dark."
    ],
    21: [  # at
        "I'll meet you at the restaurant at seven o'clock.",
        "She's very good at solving complex mathematical problems.",
        "The train arrives at the station in fifteen minutes."
    ],
    22: [  # by
        "The novel was written by one of my favorite authors.",
        "Please submit your assignment by Friday afternoon.",
        "They traveled to the island by ferry across the bay."
    ],
    23: [  # have
        "I have three younger sisters and one older brother.",
        "We have to leave now or we'll miss the train.",
        "They have lived in this neighborhood for twenty years."
    ],
    24: [  # which
        "I can't decide which dress to wear to the party.",
        "The museum, which opened last year, attracts thousands of visitors.",
        "She asked me which route would be faster to downtown."
    ],
    25: [  # or
        "Would you prefer tea or coffee with your breakfast?",
        "You can pay by credit card or cash at the register.",
        "We could go to the movies or stay home and watch something."
    ],
    26: [  # from
        "She received a beautiful postcard from her friend in Italy.",
        "The train from London arrives at platform five.",
        "I learned so much from that challenging experience."
    ],
    27: [  # this
        "This is the best chocolate cake I've ever tasted.",
        "I bought this jacket last week during the big sale.",
        "Can you explain this problem to me one more time?"
    ],
    28: [  # him
        "Tell him that I'll call back later this evening.",
        "The award was given to him for his outstanding research.",
        "I saw him walking his dog in the park yesterday."
    ],
    29: [  # but
        "The exam was difficult, but I think I did well.",
        "She wanted to go, but she had too much work to finish.",
        "It's expensive, but the quality is worth the extra cost."
    ],
    30: [  # all
        "All students must complete the assignment by Monday.",
        "She spent all day cleaning and organizing her apartment.",
        "They all agreed that the presentation was excellent."
    ],
    31: [  # she
        "She graduated from university with honors last spring.",
        "When she smiled, everyone in the room felt happier.",
        "I think she is the perfect candidate for this position."
    ]
}

if __name__ == "__main__":
    import openpyxl

    print("Updating Excel file with LLM-generated sentences...")
    wb = openpyxl.load_workbook('MyEnglishWords.xlsx')
    ws = wb.active

    count = 0
    for row_num, sents in sentences.items():
        ws.cell(row=row_num, column=9).value = sents[0]
        ws.cell(row=row_num, column=10).value = sents[1]
        ws.cell(row=row_num, column=11).value = sents[2]
        count += 1

    wb.save('MyEnglishWords.xlsx')
    print(f"✓ Updated {count} words with {count * 3} new sentences!")
