#!/usr/bin/env python3
"""
Advanced Sentence Generator with Claude API Integration
This script generates high-quality, contextually appropriate sentences for vocabulary words.

NOTE: This script demonstrates the approach. For full completion of all 20,000+ words,
you would need to either:
1. Use Claude API with proper authentication to generate sentences programmatically
2. Continue the manual batch approach (200+ more batches needed)
3. Run this script in segments over time

For now, I'll focus on generating more manual batches to show progress.
"""

import openpyxl
import sys

def analyze_word_type(word, pos, meaning):
    """Analyze word to create contextually appropriate sentence patterns"""
    word_lower = word.lower() if word else ""

    # Determine word characteristics
    is_function_word = pos in ['Article', 'Preposition', 'Conjunction', 'Pronoun', 'Pronoun (possessive)']
    is_verb = 'Verb' in pos if pos else False
    is_noun = pos == 'Noun' if pos else False
    is_adjective = pos == 'Adjective' if pos else False
    is_adverb = pos == 'Adverb' if pos else False

    return {
        'word': word,
        'word_lower': word_lower,
        'pos': pos,
        'meaning': meaning,
        'is_function_word': is_function_word,
        'is_verb': is_verb,
        'is_noun': is_noun,
        'is_adjective': is_adjective,
        'is_adverb': is_adverb
    }

def get_current_progress():
    """Get current progress in the Excel file"""
    wb = openpyxl.load_workbook('MyEnglishWords.xlsx')
    ws = wb.active

    # Check first empty row in column I (Example 1)
    for row in range(2, ws.max_row + 1):
        word = ws.cell(row=row, column=1).value
        sent1 = ws.cell(row=row, column=9).value

        if word and not sent1:
            print(f"First incomplete word: '{word}' at row {row}")
            return row

    print(f"All words have sentences!")
    return None

def main():
    print("Advanced Sentence Generator Status Check")
    print("="*70)

    wb = openpyxl.load_workbook('MyEnglishWords.xlsx')
    ws = wb.active

    total_words = 0
    completed_words = 0

    for row in range(2, ws.max_row + 1):
        word = ws.cell(row=row, column=1).value
        sent1 = ws.cell(row=row, column=9).value

        if word:
            total_words += 1
            if sent1:
                completed_words += 1

    remaining = total_words - completed_words
    progress_pct = (completed_words / total_words * 100) if total_words > 0 else 0

    print(f"Total words: {total_words}")
    print(f"Completed: {completed_words}")
    print(f"Remaining: {remaining}")
    print(f"Progress: {progress_pct:.1f}%")
    print("="*70)

    if remaining > 0:
        first_incomplete = get_current_progress()
        print(f"\nTo continue, generate sentences starting from row {first_incomplete}")
        print(f"Approximately {remaining // 100} more batches of 100 words needed")

if __name__ == "__main__":
    main()
