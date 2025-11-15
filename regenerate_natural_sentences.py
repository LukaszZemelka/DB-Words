#!/usr/bin/env python3
"""
COMPREHENSIVE NATURAL SENTENCE GENERATOR
Every sentence is uniquely crafted to demonstrate authentic English usage.
NO TEMPLATES - each sentence is written individually for educational value.
"""

import openpyxl

# ===== AUTHENTIC, HIGH-QUALITY SENTENCES FOR EACH WORD =====
# Each word gets 3 completely unique, natural sentences that demonstrate real usage

NATURAL_SENTENCES = {
    "inanity": [
        "The inanity of arguing about such trivial matters frustrated everyone at the meeting.",
        "She couldn't bear to listen to another moment of his mindless inanity.",
        "The TV show's complete inanity made me wonder why anyone would watch it."
    ],
    "Jurassic": [
        "The Jurassic period lasted from about 200 to 145 million years ago.",
        "Paleontologists discovered fascinating Jurassic fossils in the limestone quarry.",
        "Dinosaurs reached their peak diversity during the Jurassic era."
    ],
    "neolithic": [
        "The neolithic revolution marked humanity's transition from hunting to agriculture.",
        "Archaeologists found neolithic tools and pottery at the ancient settlement.",
        "During the neolithic period, people began building permanent villages."
    ],
    "rafter": [
        "The barn's old wooden rafters creaked loudly during strong winds.",
        "Birds had built their nests high up in the cathedral's rafters.",
        "The carpenter measured each rafter carefully before installation."
    ],
    "saturnalia": [
        "Ancient Romans celebrated Saturnalia with feasting and gift-giving in December.",
        "The festival descended into a saturnalia of drunken revelry.",
        "During Saturnalia, social norms were temporarily reversed."
    ],
    "silt": [
        "The river deposited layers of fine silt along its banks each spring.",
        "Farmers valued the nutrient-rich silt left by seasonal flooding.",
        "The harbor required constant dredging to remove accumulated silt."
    ],
    "uproot": [
        "The hurricane was strong enough to uproot massive oak trees.",
        "War forced millions of families to uproot themselves and flee.",
        "She didn't want to uproot her children by moving to another city."
    ],
    "caw": [
        "The harsh caw of crows echoed across the empty field.",
        "A raven began to caw loudly from the top of the fence post.",
        "I woke to the sound of seagulls cawing outside my window."
    ],
    "anymore": [
        "I don't live in Boston anymore; I moved to Seattle last year.",
        "She doesn't call me anymore since we had that argument.",
        "Do you even care about this project anymore?"
    ],
    "apportion": [
        "The committee will apportion the budget fairly among all departments.",
        "It's difficult to apportion blame when everyone shares responsibility.",
        "The judge must apportion damages based on each party's liability."
    ],
    "coif": [
        "Medieval women often wore a simple linen coif to cover their hair.",
        "The hairstylist created an elegant coif for the bride's wedding day.",
        "Nuns traditionally wore a white coif as part of their habit."
    ],
    "denominational": [
        "The school welcomes students from various denominational backgrounds.",
        "Denominational differences sometimes create tensions within communities.",
        "They founded a non-denominational church that serves all Christians."
    ],
    "eyne": [
        "In archaic English poetry, 'eyne' means eyes.",
        "Shakespeare occasionally used 'eyne' as the plural of 'eye' in his sonnets.",
        "The old ballad speaks of 'beauteous eyne' shining bright."
    ],
    "fertilization": [
        "After fertilization occurs, the egg begins to divide rapidly.",
        "Bees play a crucial role in the fertilization of flowering plants.",
        "In vitro fertilization has helped many couples conceive children."
    ],
    "gunnery": [
        "The naval officer specialized in gunnery and ballistics.",
        "Accurate gunnery requires extensive training and practice.",
        "He earned commendations for his exceptional gunnery skills during the war."
    ],
    "infuriate": [
        "His constant lateness never failed to infuriate his punctual colleagues.",
        "The airline's poor customer service will infuriate even the most patient travelers.",
        "Nothing could infuriate her more than people who break their promises."
    ],
    "murk": [
        "The submarine descended slowly into the murk of the deep ocean.",
        "Through the murk of the foggy morning, we could barely see the road.",
        "Her past remained hidden in the murk of incomplete records."
    ],
    "peristyle": [
        "The ancient temple featured a magnificent peristyle surrounding the inner sanctuary.",
        "Visitors admired the marble columns of the Roman peristyle courtyard.",
        "Greek architecture often incorporated a peristyle into public buildings."
    ],
    "unimpeded": [
        "The runner had an unimpeded view of the finish line ahead.",
        "Traffic flowed unimpeded along the newly constructed highway.",
        "We need unimpeded access to all company records for the audit."
    ],
    "warren": [
        "The rabbit warren extended underground in a complex network of tunnels.",
        "They lived in a warren of narrow, winding streets in the old quarter.",
        "The apartment building was a warren of tiny studios and hallways."
    ],
    "bivalve": [
        "Clams, oysters, and mussels are all types of bivalve mollusks.",
        "The bivalve uses its two hinged shells for protection from predators.",
        "Marine biologists study how bivalve filter-feeders improve water quality."
    ],
    "Cracow": [
        "Cracow is one of Poland's oldest and most beautiful cities.",
        "The University of Cracow was founded in 1364.",
        "Tourists flock to Cracow to see its medieval architecture and rich history."
    ],
    "curdle": [
        "Adding lemon juice will cause the milk to curdle immediately.",
        "His blood seemed to curdle at the horrifying sight before him.",
        "The cream curdled because it was left out in the heat too long."
    ],
    "curfew": [
        "The city imposed a strict curfew from 10 PM to 6 AM.",
        "Teenagers must obey their parents' curfew or face consequences.",
        "During wartime, violating curfew could result in arrest."
    ],
    "experimentation": [
        "Scientific progress depends on careful experimentation and observation.",
        "The artist's work shows bold experimentation with color and form.",
        "Animal rights activists oppose experimentation on living creatures."
    ],
    "NZ": [
        "NZ is the common abbreviation for New Zealand.",
        "I'm planning a trip to NZ next summer to see the mountains.",
        "Many films have been shot in NZ because of its stunning landscapes."
    ],
    "RI": [
        "RI stands for Rhode Island, the smallest US state.",
        "She graduated from Brown University in Providence, RI.",
        "Despite its size, RI has a rich colonial history."
    ],
    "seraphim": [
        "In religious art, seraphim are depicted with six wings.",
        "The seraphim are described as the highest order of angels.",
        "Medieval paintings often show seraphim surrounding God's throne."
    ],
    "somnambulist": [
        "The somnambulist walked through the house while completely asleep.",
        "Doctors warned that waking a somnambulist suddenly could be dangerous.",
        "She became a somnambulist after starting her new medication."
    ],
    "wainscoting": [
        "The dining room featured elegant oak wainscoting along the walls.",
        "Installing wainscoting can add character to a plain room.",
        "The Victorian house still had its original wooden wainscoting."
    ],
    "wattle": [
        "The turkey's red wattle shook as it strutted across the yard.",
        "Aboriginal Australians built shelters using wattle and daub construction.",
        "The bright yellow wattle blooms in Australian springtime."
    ],
    "griffin": [
        "The griffin is a mythical creature with an eagle's head and lion's body.",
        "Stone griffins guarded the entrance to the ancient palace.",
        "In heraldry, the griffin symbolizes courage and vigilance."
    ],
    "internecine": [
        "The internecine conflict within the party weakened their electoral chances.",
        "Years of internecine warfare devastated the small nation.",
        "Corporate internecine struggles damaged the company's reputation."
    ],
    "layout": [
        "The magazine's layout was clean, modern, and easy to read.",
        "Before construction begins, architects finalize the building's layout.",
        "I don't like the awkward layout of this apartment's rooms."
    ],
    "misnomer": [
        "Calling it a 'parking lot' is a misnomer since no parking is allowed.",
        "The term 'jellyfish' is actually a misnomer because they're not fish.",
        "To call him lazy would be a complete misnomer; he works constantly."
    ],
    "objector": [
        "He registered as a conscientious objector and refused military service.",
        "The court recognized her status as a religious objector to vaccination.",
        "Not a single objector raised concerns during the public hearing."
    ],
    "unsigned": [
        "The contract remained unsigned despite weeks of negotiations.",
        "Police found an unsigned suicide note at the scene.",
        "The unsigned painting was later attributed to Rembrandt."
    ],
    "castanet": [
        "Flamenco dancers click wooden castanets rhythmically while performing.",
        "She learned to play the castanet during her year in Spain.",
        "The sharp clicking of castanets added authentic flavor to the music."
    ],
    "CR": [
        "CR is the abbreviation for Costa Rica in international codes.",
        "The document's CR symbol indicates a carriage return.",
        "Our company uses CR to denote credit in accounting records."
    ],
    "maroon": [
        "Pirates sometimes marooned their enemies on deserted islands.",
        "She wore a beautiful maroon dress to the evening event.",
        "The school's colors are maroon and white."
    ],
    "polyglot": [
        "As a polyglot, she speaks seven languages fluently.",
        "The polyglot community in this neighborhood speaks dozens of languages.",
        "His polyglot abilities made him invaluable as a diplomat."
    ],
    "solecism": [
        "Using 'ain't' in formal writing is considered a grammatical solecism.",
        "The professor winced at the student's obvious solecism.",
        "Saying 'between you and I' is a common solecism."
    ],
    "cayuse": [
        "Cowboys in the Old West often rode tough, hardy cayuse ponies.",
        "The cayuse was originally bred by Native American tribes.",
        "His small cayuse could traverse rough mountain terrain easily."
    ],
    "Hecate": [
        "In Greek mythology, Hecate was the goddess of witchcraft and crossroads.",
        "Ancient Greeks often left offerings for Hecate at three-way intersections.",
        "Shakespeare references Hecate in his play Macbeth."
    ],
    "homosexual": [
        "Society's attitudes toward homosexual relationships have evolved significantly.",
        "The organization fights discrimination against homosexual individuals.",
        "Many countries now legally recognize homosexual marriage."
    ],
    "Jackie": [
        "Jackie Kennedy was known for her elegance and style.",
        "My friend Jackie is throwing a party next weekend.",
        "Jackie Robinson broke baseball's color barrier in 1947."
    ],
    "membranous": [
        "The membranous tissue protects the delicate internal organs.",
        "Bats have thin, membranous wings that enable flight.",
        "The doctor noticed a membranous growth during the examination."
    ],
    "respondent": [
        "The survey respondent answered all questions honestly and thoroughly.",
        "In a lawsuit, the respondent must file an answer to the complaint.",
        "Only half the respondents indicated they would support the measure."
    ],
    "Salerno": [
        "The medieval medical school at Salerno was renowned throughout Europe.",
        "Allied forces landed at Salerno during the Italian campaign in 1943.",
        "We spent a wonderful week exploring the coast near Salerno."
    ],
    "womanliness": [
        "Victorian society had rigid expectations regarding proper womanliness.",
        "She possessed a quiet strength that didn't diminish her womanliness.",
        "The traditional definition of womanliness has changed dramatically."
    ],
}


def main():
    """Apply truly natural sentences to the Excel file."""
    
    print("="*80)
    print("REGENERATING WITH TRULY NATURAL, EDUCATIONAL SENTENCES")
    print("NO TEMPLATES - Each sentence is uniquely crafted")
    print("="*80)
    
    print("\nLoading MyEnglishWords.xlsx...")
    wb = openpyxl.load_workbook('MyEnglishWords.xlsx')
    ws = wb.active
    print("✓ Loaded")
    
    print(f"\nApplying {len(NATURAL_SENTENCES)} sets of natural sentences...")
    updated = 0
    
    for row in range(16897, 20052):
        if row > ws.max_row:
            break
            
        word = ws.cell(row=row, column=1).value
        if word and word in NATURAL_SENTENCES:
            sents = NATURAL_SENTENCES[word]
            ws.cell(row=row, column=9).value = sents[0]
            ws.cell(row=row, column=10).value = sents[1]
            ws.cell(row=row, column=11).value = sents[2]
            updated += 1
    
    print(f"✓ Updated {updated} words with truly natural sentences")
    
    print("\nSaving file...")
    wb.save('MyEnglishWords.xlsx')
    print("✓ Saved successfully!")
    
    print("\n" + "="*80)
    print(f"PHASE 1 COMPLETE: {updated} words now have authentic sentences")
    print(f"NOTE: This is batch 1 of ~50. Continue adding more words...")
    print("="*80)


if __name__ == "__main__":
    main()
